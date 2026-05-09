"""
XLSX format reader functions.
"""

from pathlib import Path
from typing import Any, Iterator, List, Optional

import pyarrow as pa

from parq.formats._common import _InputMetadata, _select_schema


def _require_openpyxl() -> Any:
    """Import openpyxl lazily and raise actionable error when unavailable."""
    try:
        import openpyxl
    except ImportError as e:
        raise ValueError(
            "XLSX support requires 'openpyxl'. Install it with: pip install 'parq-cli[xlsx]'"
        ) from e
    return openpyxl


def _resolve_xlsx_sheet(workbook: Any, sheet: Optional[str]) -> Any:
    """Return the correct worksheet from an openpyxl workbook.

    sheet=None  → active sheet (original behaviour)
    sheet="2"   → workbook.worksheets[2]  (0-based integer index)
    sheet="foo" → workbook["foo"]          (sheet name)
    """
    if sheet is None:
        return workbook.active
    if sheet.isdigit():
        idx = int(sheet)
        sheets = workbook.worksheets
        if idx >= len(sheets):
            raise ValueError(
                f"Sheet index {idx} out of range — workbook has {len(sheets)} sheet(s)"
            )
        return sheets[idx]
    if sheet not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet}' not found — available sheets: {workbook.sheetnames}")
    return workbook[sheet]


def _normalize_excel_headers(header_row: List[Any]) -> List[str]:
    """Normalize empty/duplicate xlsx header names for Arrow table construction."""
    seen = set()
    normalized = []
    for idx, value in enumerate(header_row, start=1):
        base_name = (
            str(value).strip() if value is not None and str(value).strip() else f"column_{idx}"
        )
        name = base_name
        suffix = 2
        while name in seen:
            name = f"{base_name}_{suffix}"
            suffix += 1
        seen.add(name)
        normalized.append(name)
    return normalized


def _infer_arrow_type(value: Any) -> Optional[pa.DataType]:
    """Infer Arrow type for a single spreadsheet value."""
    if value is None:
        return None
    try:
        return pa.array([value]).type
    except (pa.ArrowInvalid, pa.ArrowTypeError):
        return pa.string()


def _merge_arrow_types(
    existing_type: Optional[pa.DataType], new_type: Optional[pa.DataType]
) -> Optional[pa.DataType]:
    """Merge two inferred Arrow types conservatively for spreadsheet scanning."""
    if new_type is None:
        return existing_type
    if existing_type is None or existing_type == pa.null():
        return new_type
    if new_type == pa.null() or existing_type == new_type:
        return existing_type
    if pa.types.is_integer(existing_type) and pa.types.is_floating(new_type):
        return new_type
    if pa.types.is_floating(existing_type) and pa.types.is_integer(new_type):
        return existing_type
    return pa.string()


def _coerce_value_for_field(value: Any, field: pa.Field) -> Any:
    """Coerce row values to match the selected Arrow field when needed."""
    if value is None:
        return None
    if pa.types.is_string(field.type) and not isinstance(value, str):
        return str(value)
    return value


def _scan_xlsx_structure(
    file_path: Path, sample_size: int = 1000, sheet: Optional[str] = None
) -> _InputMetadata:
    """Scan xlsx headers/schema, sampling only the first rows for type inference."""
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = _resolve_xlsx_sheet(workbook, sheet)
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return _InputMetadata(headers=tuple(), schema=pa.schema([]), num_rows=0)

        headers = _normalize_excel_headers(list(header))
        inferred_types: list[Optional[pa.DataType]] = [None] * len(headers)
        sampled_rows = 0

        for row in rows:
            sampled_rows += 1
            row_values = list(row) if row is not None else []
            for idx, _name in enumerate(headers):
                value = row_values[idx] if idx < len(row_values) else None
                inferred_types[idx] = _merge_arrow_types(
                    inferred_types[idx], _infer_arrow_type(value)
                )
            if sampled_rows >= sample_size:
                break

        schema = pa.schema(
            [
                pa.field(name, inferred_type or pa.null(), nullable=True)
                for name, inferred_type in zip(headers, inferred_types)
            ]
        )
        return _InputMetadata(
            headers=tuple(headers),
            schema=schema,
            num_rows=sampled_rows if sampled_rows < sample_size else None,
        )
    finally:
        workbook.close()


def _count_xlsx_rows(file_path: Path, sheet: Optional[str] = None) -> int:
    """Count xlsx data rows exactly, excluding the header row."""
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = _resolve_xlsx_sheet(workbook, sheet)
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return 0
        return sum(1 for _ in rows)
    finally:
        workbook.close()


def _iter_xlsx_batches(
    file_path: Path,
    metadata: _InputMetadata,
    columns: Optional[List[str]] = None,
    batch_size: int = 1000,
    sheet: Optional[str] = None,
) -> Iterator[pa.RecordBatch]:
    """Stream xlsx input as Arrow record batches."""
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = _resolve_xlsx_sheet(workbook, sheet)
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return

        selected_columns = list(metadata.headers) if columns is None else columns
        selected_indices = [metadata.headers.index(column) for column in selected_columns]
        selected_schema = _select_schema(metadata.schema, selected_columns)
        selected_fields = [selected_schema.field(column) for column in selected_columns]
        buffered_rows: list[dict[str, Any]] = []

        for row in rows:
            row_values = list(row) if row is not None else []
            buffered_rows.append(
                {
                    column: _coerce_value_for_field(
                        row_values[idx] if idx < len(row_values) else None,
                        field,
                    )
                    for column, idx, field in zip(
                        selected_columns, selected_indices, selected_fields
                    )
                }
            )
            if len(buffered_rows) >= batch_size:
                yield pa.Table.from_pylist(buffered_rows, schema=selected_schema).to_batches()[0]
                buffered_rows = []

        if buffered_rows:
            yield pa.Table.from_pylist(buffered_rows, schema=selected_schema).to_batches()[0]
    finally:
        workbook.close()
