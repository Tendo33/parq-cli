"""
Tabular file reader module.
Provides functionality to read and inspect parquet/csv/xlsx files.
"""

from collections import deque
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Callable, Iterable, Iterator, List, Optional, Union

import pyarrow as pa
import pyarrow.csv as pacsv
import pyarrow.compute as pc
import pyarrow.parquet as pq

# Heuristic multiplier for determining when to use fast path vs optimized path
# Files with <= n * FAST_PATH_ROWS_MULTIPLIER rows use full read + slice (simpler)
# Larger files use row group optimization to reduce I/O
FAST_PATH_ROWS_MULTIPLIER = 10
SUPPORTED_INPUT_FORMATS = {".parquet", ".csv", ".xlsx"}


@dataclass
class _InputMetadata:
    """Cached input metadata for lazily-scanned non-parquet sources."""

    headers: tuple[str, ...]
    schema: pa.Schema
    num_rows: Optional[int] = None


def _validate_preview_params(n: int, schema: pa.Schema, columns: Optional[List[str]]) -> None:
    """Validate preview parameters shared by multiple readers."""
    if n < 0:
        raise ValueError(f"n must be non-negative, got {n}")

    if columns is not None:
        if len(columns) == 0:
            raise ValueError("columns cannot be empty")
        missing = [c for c in columns if c not in schema.names]
        if missing:
            raise ValueError(f"Columns not found in schema: {missing}")


def _create_empty_table(schema: pa.Schema, columns: Optional[List[str]] = None) -> pa.Table:
    """Create an empty table with source schema or selected columns."""
    selected_schema = _select_schema(schema, columns)
    fields = list(selected_schema)
    arrays = [pa.array([], type=field.type) for field in fields]
    names = [field.name for field in fields]
    return pa.Table.from_arrays(arrays, names=names)


def _select_schema(schema: pa.Schema, columns: Optional[List[str]] = None) -> pa.Schema:
    """Return full schema or selected-column schema."""
    if columns is None:
        return schema
    return pa.schema([schema.field(column) for column in columns])


def _table_schema_info(schema: pa.Schema) -> List[dict]:
    """Convert Arrow schema to formatter-compatible schema info structure."""
    return [
        {
            "name": field.name,
            "type": str(field.type),
            "nullable": field.nullable,
        }
        for field in schema
    ]


def _require_openpyxl() -> Any:
    """Import openpyxl lazily and raise actionable error when unavailable."""
    try:
        import openpyxl
    except ImportError as e:
        raise ValueError(
            "XLSX support requires 'openpyxl'. Install it with: pip install 'parq-cli[xlsx]'"
        ) from e
    return openpyxl


def _normalize_excel_headers(header_row: List[Any]) -> List[str]:
    """Normalize empty/duplicate xlsx header names for Arrow table construction."""
    seen = set()
    normalized = []
    for idx, value in enumerate(header_row, start=1):
        base_name = str(value).strip() if value is not None and str(value).strip() else f"column_{idx}"
        name = base_name
        suffix = 2
        while name in seen:
            name = f"{base_name}_{suffix}"
            suffix += 1
        seen.add(name)
        normalized.append(name)
    return normalized


def _infer_arrow_type(value: Any) -> Optional[pa.DataType]:
    """Infer Arrow type for a single spreadsheet value."""
    if value is None:
        return None
    try:
        return pa.array([value]).type
    except (pa.ArrowInvalid, pa.ArrowTypeError):
        return pa.string()


def _merge_arrow_types(
    existing_type: Optional[pa.DataType], new_type: Optional[pa.DataType]
) -> Optional[pa.DataType]:
    """Merge two inferred Arrow types conservatively for spreadsheet scanning."""
    if new_type is None:
        return existing_type
    if existing_type is None or existing_type == pa.null():
        return new_type
    if new_type == pa.null() or existing_type == new_type:
        return existing_type
    if pa.types.is_integer(existing_type) and pa.types.is_floating(new_type):
        return new_type
    if pa.types.is_floating(existing_type) and pa.types.is_integer(new_type):
        return existing_type
    return pa.string()


def _coerce_value_for_field(value: Any, field: pa.Field) -> Any:
    """Coerce row values to match the selected Arrow field when needed."""
    if value is None:
        return None
    if pa.types.is_string(field.type) and not isinstance(value, str):
        return str(value)
    return value


def _iter_csv_batches(
    file_path: Path,
    columns: Optional[List[str]] = None,
) -> Iterator[pa.RecordBatch]:
    """Stream CSV input as Arrow record batches."""
    convert_options = None
    if columns is not None:
        convert_options = pacsv.ConvertOptions(include_columns=columns)
    yield from pacsv.open_csv(file_path, convert_options=convert_options)


def _scan_csv_metadata(file_path: Path, include_row_count: bool = False) -> _InputMetadata:
    """Scan CSV schema eagerly and count rows only when required."""
    reader = pacsv.open_csv(file_path)
    schema = reader.schema
    num_rows = sum(len(batch) for batch in reader) if include_row_count else None
    return _InputMetadata(headers=tuple(schema.names), schema=schema, num_rows=num_rows)


def _scan_xlsx_structure(file_path: Path, sample_size: int = 1000) -> _InputMetadata:
    """Scan xlsx headers/schema, sampling only the first rows for type inference."""
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return _InputMetadata(headers=tuple(), schema=pa.schema([]), num_rows=0)

        headers = _normalize_excel_headers(list(header))
        inferred_types: list[Optional[pa.DataType]] = [None] * len(headers)
        sampled_rows = 0

        for row in rows:
            sampled_rows += 1
            row_values = list(row) if row is not None else []
            for idx, _name in enumerate(headers):
                value = row_values[idx] if idx < len(row_values) else None
                inferred_types[idx] = _merge_arrow_types(
                    inferred_types[idx], _infer_arrow_type(value)
                )
            if sampled_rows >= sample_size:
                break

        schema = pa.schema(
            [
                pa.field(name, inferred_type or pa.null(), nullable=True)
                for name, inferred_type in zip(headers, inferred_types)
            ]
        )
        return _InputMetadata(
            headers=tuple(headers),
            schema=schema,
            num_rows=sampled_rows if sampled_rows < sample_size else None,
        )
    finally:
        workbook.close()


def _count_xlsx_rows(file_path: Path) -> int:
    """Count xlsx data rows exactly, excluding the header row."""
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return 0
        return sum(1 for _ in rows)
    finally:
        workbook.close()


def _iter_xlsx_batches(
    file_path: Path,
    metadata: _InputMetadata,
    columns: Optional[List[str]] = None,
    batch_size: int = 1000,
) -> Iterator[pa.RecordBatch]:
    """Stream xlsx input as Arrow record batches."""
    openpyxl = _require_openpyxl()
    workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return

        selected_columns = list(metadata.headers) if columns is None else columns
        selected_indices = [metadata.headers.index(column) for column in selected_columns]
        selected_schema = _select_schema(metadata.schema, selected_columns)
        selected_fields = [selected_schema.field(column) for column in selected_columns]
        buffered_rows: list[dict[str, Any]] = []

        for row in rows:
            row_values = list(row) if row is not None else []
            buffered_rows.append(
                {
                    column: _coerce_value_for_field(
                        row_values[idx] if idx < len(row_values) else None,
                        field,
                    )
                    for column, idx, field in zip(selected_columns, selected_indices, selected_fields)
                }
            )
            if len(buffered_rows) >= batch_size:
                yield pa.Table.from_pylist(buffered_rows, schema=selected_schema).to_batches()[0]
                buffered_rows = []

        if buffered_rows:
            yield pa.Table.from_pylist(buffered_rows, schema=selected_schema).to_batches()[0]
    finally:
        workbook.close()


def _collect_head_from_batches(
    batches: Iterable[pa.RecordBatch],
    n: int,
    schema: pa.Schema,
) -> pa.Table:
    """Collect first n rows from a batch iterator."""
    if n == 0:
        return _create_empty_table(schema)

    collected_batches: list[pa.RecordBatch] = []
    rows_collected = 0
    for batch in batches:
        if rows_collected >= n:
            break
        rows_to_take = min(n - rows_collected, len(batch))
        collected_batches.append(batch if rows_to_take == len(batch) else batch.slice(0, rows_to_take))
        rows_collected += rows_to_take
        if rows_collected >= n:
            break

    if not collected_batches:
        return _create_empty_table(schema)
    return pa.Table.from_batches(collected_batches, schema=schema)


def _collect_tail_from_batches(
    batches: Iterable[pa.RecordBatch],
    n: int,
    schema: pa.Schema,
) -> pa.Table:
    """Collect last n rows using per-column ring buffers instead of row dicts."""
    if n == 0:
        return _create_empty_table(schema)

    tail_columns = {field.name: deque(maxlen=n) for field in schema}
    for batch in batches:
        batch_dict = batch.to_pydict()
        for name in schema.names:
            tail_columns[name].extend(batch_dict[name])

    if not tail_columns or not any(tail_columns.values()):
        return _create_empty_table(schema)
    arrays = [
        pa.array(list(tail_columns[field.name]), type=field.type) for field in schema
    ]
    return pa.Table.from_arrays(arrays, schema=schema)


def _collect_preview_from_batches(
    batches: Iterable[pa.RecordBatch],
    n: int,
    schema: pa.Schema,
    from_tail: bool = False,
) -> pa.Table:
    """Collect a preview window from a batch iterator without materializing the full input."""
    if from_tail:
        return _collect_tail_from_batches(batches, n, schema)
    return _collect_head_from_batches(batches, n, schema)


def _resolve_split_shape(
    total_rows: int,
    file_count: Optional[int] = None,
    record_count: Optional[int] = None,
) -> List[int]:
    """Resolve split chunk sizes with existing validation semantics."""
    if file_count is None and record_count is None:
        raise ValueError("Either file_count or record_count must be specified")
    if file_count is not None and record_count is not None:
        raise ValueError("file_count and record_count are mutually exclusive")
    if total_rows == 0:
        raise ValueError("Cannot split empty file")

    if file_count is not None:
        if file_count <= 0:
            raise ValueError("file_count must be positive")
        if file_count > total_rows:
            raise ValueError("file_count cannot exceed total rows")
        base_rows = total_rows // file_count
        remainder = total_rows % file_count
        return [base_rows + (1 if i < remainder else 0) for i in range(file_count)]

    assert record_count is not None
    if record_count <= 0:
        raise ValueError("record_count must be positive")
    full_chunks, remainder = divmod(total_rows, record_count)
    chunk_sizes = [record_count] * full_chunks
    if remainder:
        chunk_sizes.append(remainder)
    return chunk_sizes


def _coerce_output_value(value: Any) -> Any:
    """Normalize nested values before writing them to text-oriented output formats."""
    if isinstance(value, (dict, list, set, tuple)):
        return str(value)
    return value


class _CsvChunkWriter:
    """Incremental CSV writer backed by PyArrow CSVWriter."""

    def __init__(self, output_path: Path, schema: pa.Schema):
        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._writer = pacsv.CSVWriter(output_path, schema)

    def write_batch(self, batch: pa.RecordBatch) -> None:
        self._writer.write(batch)

    def close(self) -> None:
        self._writer.close()


class _XlsxChunkWriter:
    """Incremental XLSX writer using openpyxl write-only mode."""

    def __init__(self, output_path: Path, schema: pa.Schema):
        openpyxl = _require_openpyxl()
        self._output_path = output_path
        self._output_path.parent.mkdir(parents=True, exist_ok=True)
        self._workbook = openpyxl.Workbook(write_only=True)
        self._worksheet = self._workbook.create_sheet()
        self._column_names = schema.names
        self._worksheet.append(self._column_names)

    def write_batch(self, batch: pa.RecordBatch) -> None:
        batch_dict = batch.to_pydict()
        for row_idx in range(len(batch)):
            row = [
                _coerce_output_value(batch_dict[col_name][row_idx]) for col_name in self._column_names
            ]
            self._worksheet.append(row)

    def close(self) -> None:
        self._workbook.save(self._output_path)
        self._workbook.close()


def _open_chunk_writer(output_path: Path, schema: pa.Schema, compression: Optional[str] = None) -> Any:
    """Open incremental writer for supported output suffixes."""
    suffix = output_path.suffix.lower()
    if suffix == ".parquet":
        output_path.parent.mkdir(parents=True, exist_ok=True)
        return pq.ParquetWriter(output_path, schema, compression=compression)
    if suffix == ".csv":
        return _CsvChunkWriter(output_path, schema)
    if suffix == ".xlsx":
        return _XlsxChunkWriter(output_path, schema)
    raise ValueError(f"Unsupported output file format: {suffix or '<none>'}")


def _resolve_output_files(output_pattern: str, num_files: int) -> List[Path]:
    """Validate output pattern and preflight all output file paths."""
    try:
        output_pattern % 0
    except (TypeError, ValueError) as e:
        raise ValueError(f"Invalid output pattern format: {e}") from e

    output_files = [Path(output_pattern % i) for i in range(num_files)]
    for output_path in output_files:
        if output_path.exists():
            raise FileExistsError(f"Output file already exists: {output_path}")
    return output_files


def _validate_output_pattern(output_pattern: str) -> None:
    """Validate a printf-style output pattern."""
    try:
        output_pattern % 0
    except (TypeError, ValueError) as e:
        raise ValueError(f"Invalid output pattern format: {e}") from e


def _open_validated_output_path(output_pattern: str, index: int) -> Path:
    """Resolve one output path from a pattern and reject existing targets."""
    output_path = Path(output_pattern % index)
    if output_path.exists():
        raise FileExistsError(f"Output file already exists: {output_path}")
    return output_path


def _split_batches_to_files(
    batches: Iterable[pa.RecordBatch],
    schema: pa.Schema,
    output_files: List[Path],
    chunk_sizes: List[int],
    total_rows: int,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    compression: Optional[str] = None,
) -> List[Path]:
    """Stream record batches into split output files with cleanup on failure."""
    created_output_paths: List[Path] = []
    current_writer: Optional[Any] = None

    try:
        current_file_idx = 0
        current_file_rows = 0
        target_rows_for_current_file = chunk_sizes[current_file_idx]
        rows_processed = 0
        current_writer = _open_chunk_writer(
            output_files[current_file_idx], schema, compression=compression
        )
        created_output_paths.append(output_files[current_file_idx])

        for batch in batches:
            batch_offset = 0
            while batch_offset < len(batch):
                rows_remaining_in_file = target_rows_for_current_file - current_file_rows
                rows_to_write = min(rows_remaining_in_file, len(batch) - batch_offset)
                slice_batch = batch if rows_to_write == len(batch) and batch_offset == 0 else batch.slice(
                    batch_offset, rows_to_write
                )
                current_writer.write_batch(slice_batch)

                batch_offset += rows_to_write
                current_file_rows += rows_to_write
                rows_processed += rows_to_write

                if progress_callback:
                    progress_callback(rows_processed, total_rows)

                if current_file_rows >= target_rows_for_current_file and current_file_idx < len(output_files) - 1:
                    current_writer.close()
                    current_file_idx += 1
                    current_file_rows = 0
                    target_rows_for_current_file = chunk_sizes[current_file_idx]
                    current_writer = _open_chunk_writer(
                        output_files[current_file_idx], schema, compression=compression
                    )
                    created_output_paths.append(output_files[current_file_idx])

        if current_writer is not None:
            current_writer.close()
            current_writer = None
    except Exception:
        if current_writer is not None:
            try:
                current_writer.close()
            except Exception:
                pass

        for output_path in created_output_paths:
            try:
                output_path.unlink()
            except FileNotFoundError:
                pass
            except OSError:
                pass
        raise

    return output_files


def _stream_split_by_record_count(
    batches: Iterable[pa.RecordBatch],
    schema: pa.Schema,
    output_pattern: str,
    record_count: int,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    compression: Optional[str] = None,
) -> tuple[List[Path], int]:
    """Split batches in a single pass when chunking by record count."""
    if record_count <= 0:
        raise ValueError("record_count must be positive")

    _validate_output_pattern(output_pattern)
    created_output_paths: List[Path] = []
    current_writer: Optional[Any] = None
    current_rows = 0
    rows_processed = 0
    current_file_idx = 0

    try:
        for batch in batches:
            batch_offset = 0
            while batch_offset < len(batch):
                if current_writer is None:
                    output_path = _open_validated_output_path(output_pattern, current_file_idx)
                    current_writer = _open_chunk_writer(output_path, schema, compression=compression)
                    created_output_paths.append(output_path)
                    current_rows = 0

                rows_remaining = record_count - current_rows
                rows_to_write = min(rows_remaining, len(batch) - batch_offset)
                slice_batch = (
                    batch
                    if rows_to_write == len(batch) and batch_offset == 0
                    else batch.slice(batch_offset, rows_to_write)
                )
                current_writer.write_batch(slice_batch)
                batch_offset += rows_to_write
                current_rows += rows_to_write
                rows_processed += rows_to_write

                if progress_callback is not None:
                    progress_callback(rows_processed, rows_processed)

                if current_rows >= record_count:
                    current_writer.close()
                    current_writer = None
                    current_file_idx += 1

        if current_writer is not None:
            current_writer.close()
            current_writer = None
    except Exception:
        if current_writer is not None:
            try:
                current_writer.close()
            except Exception:
                pass
        for output_path in created_output_paths:
            try:
                output_path.unlink()
            except FileNotFoundError:
                pass
            except OSError:
                pass
        raise

    if rows_processed == 0:
        raise ValueError("Cannot split empty file")
    return created_output_paths, rows_processed


def _write_batches_to_output(
    batches: Iterable[pa.RecordBatch],
    schema: pa.Schema,
    output_path: Path,
    compression: Optional[str] = None,
) -> int:
    """Write a batch iterator to a single output file."""
    if output_path.exists():
        raise FileExistsError(f"Output file already exists: {output_path}")

    writer = _open_chunk_writer(output_path, schema, compression=compression)
    rows_written = 0
    try:
        for batch in batches:
            writer.write_batch(batch)
            rows_written += len(batch)
    except Exception:
        try:
            writer.close()
        except Exception:
            pass
        try:
            output_path.unlink()
        except FileNotFoundError:
            pass
        except OSError:
            pass
        raise

    writer.close()
    return rows_written


def _compute_table_stats(table: pa.Table, limit: int = 50) -> List[dict]:
    """Compute column statistics for a materialized Arrow table."""
    stats_rows: List[dict] = []
    for field in table.schema:
        column = table[field.name]
        null_count = int(column.null_count)
        count = len(column) - null_count
        row = {
            "name": field.name,
            "type": str(field.type),
            "count": count,
            "null_count": null_count,
            "min": None,
            "max": None,
            "mean": None,
        }
        if count > 0 and (
            pa.types.is_integer(field.type)
            or pa.types.is_floating(field.type)
            or pa.types.is_decimal(field.type)
        ):
            row["min"] = pc.min(column).as_py()
            row["max"] = pc.max(column).as_py()
            mean_value = pc.mean(column)
            row["mean"] = None if mean_value is None else mean_value.as_py()
        stats_rows.append(row)
        if len(stats_rows) >= limit:
            break
    return stats_rows


def _compute_stats_from_batches(
    batches: Iterable[pa.RecordBatch],
    schema: pa.Schema,
    limit: int = 50,
) -> List[dict]:
    """Compute column stats incrementally from batches without full materialization."""
    tracked_fields = list(schema)[:limit]
    stats: dict[str, dict[str, Any]] = {
        field.name: {
            "name": field.name,
            "type": str(field.type),
            "count": 0,
            "null_count": 0,
            "min": None,
            "max": None,
            "_sum": 0,
            "mean": None,
            "_numeric": (
                pa.types.is_integer(field.type)
                or pa.types.is_floating(field.type)
                or pa.types.is_decimal(field.type)
            ),
        }
        for field in tracked_fields
    }

    for batch in batches:
        for index, field in enumerate(tracked_fields):
            array = batch.column(index)
            state = stats[field.name]
            state["null_count"] += array.null_count
            state["count"] += len(array) - array.null_count
            if not state["_numeric"] or len(array) == array.null_count:
                continue

            batch_min = pc.min(array).as_py()
            batch_max = pc.max(array).as_py()
            batch_sum = pc.sum(array).as_py()

            if state["min"] is None or (batch_min is not None and batch_min < state["min"]):
                state["min"] = batch_min
            if state["max"] is None or (batch_max is not None and batch_max > state["max"]):
                state["max"] = batch_max
            if batch_sum is not None:
                state["_sum"] += batch_sum

    results = []
    for field in tracked_fields:
        state = stats[field.name]
        if state["_numeric"] and state["count"] > 0:
            state["mean"] = state["_sum"] / state["count"]
        del state["_sum"]
        del state["_numeric"]
        results.append(state)
    return results


class ParquetReader:
    """Parquet file reader with metadata inspection capabilities."""

    def __init__(self, file_path: str):
        """
        Initialize ParquetReader with a file path.

        Args:
            file_path: Path to the Parquet file
        """
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        self._parquet_file = pq.ParquetFile(self.file_path)
        self.last_split_total_rows: Optional[int] = None
        self.last_write_total_rows: Optional[int] = None

    @property
    def metadata(self) -> pq.FileMetaData:
        """Get file metadata."""
        return self._parquet_file.metadata

    @property
    def schema(self) -> pa.Schema:
        """Get file schema."""
        return self._parquet_file.schema_arrow

    @property
    def num_rows(self) -> int:
        """Get total number of rows."""
        return self.metadata.num_rows

    @property
    def num_columns(self) -> int:
        """Get total number of columns (logical schema columns)."""
        return len(self.schema)

    @property
    def num_physical_columns(self) -> int:
        """Get total number of physical columns (from metadata)."""
        return self.metadata.num_columns

    @property
    def num_row_groups(self) -> int:
        """Get number of row groups."""
        return self.metadata.num_row_groups

    def get_metadata_dict(self, fast: bool = False) -> dict:
        """
        Get metadata as a dictionary.

        Returns:
            Dictionary containing file metadata
        """
        metadata_dict = {
            "file_path": str(self.file_path),
            "num_columns": self.num_columns,
            "file_size": self.file_path.stat().st_size,
            "num_row_groups": self.num_row_groups,
            "format_version": self.metadata.format_version,
        }

        if not fast:
            metadata_dict["num_rows"] = self.num_rows

        # Add physical columns right after logical columns if different
        if self.num_physical_columns != self.num_columns:
            metadata_dict["num_physical_columns"] = self.num_physical_columns

        if not fast:
            compression = self._get_compression_summary()
            if compression is not None:
                metadata_dict["compression"] = compression
            metadata_dict.update(
                {
                    "serialized_size": self.metadata.serialized_size,
                    "created_by": self.metadata.created_by,
                }
            )

        return metadata_dict

    def get_schema_info(self) -> List[dict]:
        """
        Get schema information as a list of column details.

        Returns:
            List of dictionaries with column information
        """
        schema_info = []
        for field in self.schema:
            schema_info.append(
                {
                    "name": field.name,
                    "type": str(field.type),
                    "nullable": field.nullable,
                }
            )
        return schema_info

    def _read_preview(self, n: int, columns: Optional[List[str]] = None, from_tail: bool = False) -> pa.Table:
        """Read a parquet preview window using shared validation and row-group planning."""
        _validate_preview_params(n, self.schema, columns)
        if n == 0:
            return _create_empty_table(self.schema, columns=columns)

        if self.num_rows <= n * FAST_PATH_ROWS_MULTIPLIER or self.num_row_groups == 1:
            table = self._parquet_file.read(columns=columns)
            start = max(0, self.num_rows - n) if from_tail else 0
            length = n if from_tail else min(n, self.num_rows)
            return table.slice(start, length)

        rows_remaining = n
        row_groups = []
        indices = range(self.num_row_groups - 1, -1, -1) if from_tail else range(self.num_row_groups)
        for row_group_index in indices:
            row_groups.append(row_group_index)
            rows_remaining -= self.metadata.row_group(row_group_index).num_rows
            if rows_remaining <= 0:
                break

        if from_tail:
            row_groups.reverse()
        table = self._parquet_file.read_row_groups(row_groups, columns=columns)
        start = max(0, len(table) - n) if from_tail else 0
        return table.slice(start, n)

    def read_head(self, n: int = 5, columns: Optional[List[str]] = None) -> pa.Table:
        """
        Read first n rows.

        Optimized to only read necessary row groups for large files,
        significantly reducing memory usage and improving performance.

        Args:
            n: Number of rows to read (must be >= 0)
            columns: Optional list of column names to read. If None, read all columns.

        Returns:
            PyArrow table with first n rows

        Raises:
            ValueError: If n < 0, or columns is empty, or columns contains invalid names
        """
        return self._read_preview(n, columns=columns)

    def read_tail(self, n: int = 5, columns: Optional[List[str]] = None) -> pa.Table:
        """
        Read last n rows.

        Optimized to only read necessary row groups from the end of the file,
        significantly reducing memory usage and improving performance for large files.

        Args:
            n: Number of rows to read (must be >= 0)
            columns: Optional list of column names to read. If None, read all columns.

        Returns:
            PyArrow table with last n rows

        Raises:
            ValueError: If n < 0, or columns is empty, or columns contains invalid names
        """
        return self._read_preview(n, columns=columns, from_tail=True)

    def read_columns(self, columns: Optional[List[str]] = None) -> pa.Table:
        """
        Read specific columns.

        Args:
            columns: List of column names to read. If None, read all columns.

        Returns:
            PyArrow table with selected columns
        """
        return self._parquet_file.read(columns=columns)

    def split_file(
        self,
        output_pattern: str,
        file_count: Optional[int] = None,
        record_count: Optional[int] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> List[Path]:
        """
        Split parquet file into multiple files.

        Args:
            output_pattern: Output file name pattern (e.g., 'result-%06d.parquet')
            file_count: Number of output files (mutually exclusive with record_count)
            record_count: Number of records per file (mutually exclusive with file_count)
            progress_callback: Optional callback function(current, total) for progress updates

        Returns:
            List of created file paths

        Raises:
            ValueError: If both or neither of file_count/record_count are provided
            IOError: If file write fails
        """
        total_rows = self.num_rows
        chunk_sizes = _resolve_split_shape(total_rows, file_count=file_count, record_count=record_count)
        output_files = _resolve_output_files(output_pattern, len(chunk_sizes))
        result = _split_batches_to_files(
            self._parquet_file.iter_batches(batch_size=min(10000, max(chunk_sizes))),
            self.schema,
            output_files,
            chunk_sizes,
            total_rows,
            progress_callback=progress_callback,
            compression=self._get_compression_type(),
        )
        self.last_split_total_rows = total_rows
        return result

    def convert_file(self, output_path: Path, columns: Optional[List[str]] = None) -> int:
        """Convert parquet input to one supported output file."""
        rows_written = _write_batches_to_output(
            self._parquet_file.iter_batches(columns=columns),
            _select_schema(self.schema, columns),
            output_path,
            compression=self._get_compression_type() if output_path.suffix.lower() == ".parquet" else None,
        )
        self.last_write_total_rows = rows_written
        return rows_written

    def get_stats(self, columns: Optional[List[str]] = None, limit: int = 50) -> List[dict]:
        """Return simple column statistics."""
        selected_schema = _select_schema(self.schema, columns)
        return _compute_stats_from_batches(
            self._parquet_file.iter_batches(columns=columns),
            selected_schema,
            limit=limit,
        )

    def _get_compression_type(self) -> str:
        """
        Get compression type from source file.

        Returns:
            Compression type string (e.g., 'SNAPPY', 'GZIP', 'NONE')
        """
        if self.num_row_groups > 0:
            compression = self.metadata.row_group(0).column(0).compression
            return compression
        return "SNAPPY"  # Default compression

    def _get_compression_summary(self) -> Optional[str]:
        """Get compression summary from all row groups and columns.

        Optimized to scan the first row group fully, then only continue
        scanning subsequent row groups if heterogeneous compression is
        detected. Homogeneous files (the common case) exit after one
        row group, making this O(num_columns) instead of
        O(num_row_groups * num_columns).
        """
        if self.num_row_groups == 0:
            return None

        compressions: set[str] = set()
        first_rg = self.metadata.row_group(0)
        for col_idx in range(first_rg.num_columns):
            compressions.add(first_rg.column(col_idx).compression)

        if len(compressions) <= 1 and self.num_row_groups > 1:
            first_rg_types = compressions.copy()
            for rg_idx in range(1, self.num_row_groups):
                rg = self.metadata.row_group(rg_idx)
                for col_idx in range(rg.num_columns):
                    comp = rg.column(col_idx).compression
                    if comp not in first_rg_types:
                        compressions.add(comp)
                if compressions != first_rg_types:
                    break

        if not compressions:
            return None

        if len(compressions) == 1:
            return compressions.pop()

        return ", ".join(sorted(compressions))

class MultiFormatReader:
    """Reader that supports parquet/csv/xlsx with a unified CLI-facing interface."""

    def __init__(self, file_path: str):
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        suffix = self.file_path.suffix.lower()
        if suffix not in SUPPORTED_INPUT_FORMATS:
            supported = ", ".join(sorted(SUPPORTED_INPUT_FORMATS))
            raise ValueError(
                f"Unsupported file format: {suffix or '<none>'}. Supported formats: {supported}"
            )

        self.input_format = suffix.lstrip(".")
        self._parquet_reader: Optional[ParquetReader] = None
        self._metadata_cache: Optional[_InputMetadata] = None
        self.last_split_total_rows: Optional[int] = None
        self.last_write_total_rows: Optional[int] = None

        if suffix == ".parquet":
            self._parquet_reader = ParquetReader(file_path)

    def _load_metadata(self, *, include_row_count: bool = False) -> _InputMetadata:
        """Load and cache metadata without materializing the full input as a table."""
        if self._parquet_reader is not None:
            raise RuntimeError("Parquet metadata should be read from ParquetReader")

        if self.input_format == "xlsx":
            if self._metadata_cache is None:
                self._metadata_cache = _scan_xlsx_structure(self.file_path)
            if include_row_count and self._metadata_cache.num_rows is None:
                self._metadata_cache = _InputMetadata(
                    headers=self._metadata_cache.headers,
                    schema=self._metadata_cache.schema,
                    num_rows=_count_xlsx_rows(self.file_path),
                )
            return self._metadata_cache

        if self._metadata_cache is None or (include_row_count and self._metadata_cache.num_rows is None):
            self._metadata_cache = _scan_csv_metadata(
                self.file_path, include_row_count=include_row_count
            )
        return self._metadata_cache

    def _iter_input_batches(self, columns: Optional[List[str]] = None) -> Iterator[pa.RecordBatch]:
        """Iterate non-parquet input in batches with format branching in one place."""
        metadata = self._load_metadata()
        if self.input_format == "csv":
            yield from _iter_csv_batches(self.file_path, columns=columns)
            return
        yield from _iter_xlsx_batches(self.file_path, metadata, columns=columns)

    def _selected_schema(self, columns: Optional[List[str]] = None) -> pa.Schema:
        """Return full or selected schema after central column validation."""
        metadata = self._load_metadata()
        _validate_preview_params(0, metadata.schema, columns)
        return _select_schema(metadata.schema, columns)

    @property
    def schema(self) -> pa.Schema:
        if self._parquet_reader is not None:
            return self._parquet_reader.schema
        return self._load_metadata().schema

    @property
    def num_rows(self) -> int:
        if self._parquet_reader is not None:
            return self._parquet_reader.num_rows
        return self._load_metadata(include_row_count=True).num_rows or 0

    @property
    def num_columns(self) -> int:
        if self._parquet_reader is not None:
            return self._parquet_reader.num_columns
        return len(self.schema)

    @property
    def num_physical_columns(self) -> int:
        if self._parquet_reader is not None:
            return self._parquet_reader.num_physical_columns
        return self.num_columns

    @property
    def num_row_groups(self) -> int:
        if self._parquet_reader is not None:
            return self._parquet_reader.num_row_groups
        return 1

    def get_metadata_dict(self, fast: bool = False) -> dict:
        if self._parquet_reader is not None:
            return self._parquet_reader.get_metadata_dict(fast=fast)

        metadata_dict = {
            "file_path": str(self.file_path),
            "input_format": self.input_format,
            "num_columns": self.num_columns,
            "file_size": self.file_path.stat().st_size,
            "num_row_groups": self.num_row_groups,
        }
        if not fast:
            metadata_dict["num_rows"] = self.num_rows
        return metadata_dict

    def get_schema_info(self) -> List[dict]:
        if self._parquet_reader is not None:
            return self._parquet_reader.get_schema_info()
        return _table_schema_info(self.schema)

    def read_head(self, n: int = 5, columns: Optional[List[str]] = None) -> pa.Table:
        if self._parquet_reader is not None:
            return self._parquet_reader.read_head(n=n, columns=columns)

        _validate_preview_params(n, self.schema, columns)
        if n == 0:
            return _create_empty_table(self.schema, columns=columns)
        return _collect_preview_from_batches(
            self._iter_input_batches(columns=columns),
            n,
            self._selected_schema(columns),
        )

    def read_tail(self, n: int = 5, columns: Optional[List[str]] = None) -> pa.Table:
        if self._parquet_reader is not None:
            return self._parquet_reader.read_tail(n=n, columns=columns)

        _validate_preview_params(n, self.schema, columns)
        if n == 0:
            return _create_empty_table(self.schema, columns=columns)
        return _collect_preview_from_batches(
            self._iter_input_batches(columns=columns),
            n,
            self._selected_schema(columns),
            from_tail=True,
        )

    def read_columns(self, columns: Optional[List[str]] = None) -> pa.Table:
        if self._parquet_reader is not None:
            return self._parquet_reader.read_columns(columns=columns)
        selected_schema = self._selected_schema(columns)
        batches = list(self._iter_input_batches(columns=columns))
        if not batches:
            return _create_empty_table(selected_schema)
        return pa.Table.from_batches(batches, schema=selected_schema)

    def split_file(
        self,
        output_pattern: str,
        file_count: Optional[int] = None,
        record_count: Optional[int] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> List[Path]:
        _validate_output_pattern(output_pattern)

        sample_output = Path(output_pattern % 0)
        output_suffix = sample_output.suffix.lower()

        if self._parquet_reader is not None and output_suffix == ".parquet":
            output_files = self._parquet_reader.split_file(
                output_pattern=output_pattern,
                file_count=file_count,
                record_count=record_count,
                progress_callback=progress_callback,
            )
            self.last_split_total_rows = self._parquet_reader.last_split_total_rows
            return output_files

        if record_count is not None and file_count is None:
            schema = self.schema if self._parquet_reader is None else self._parquet_reader.schema
            batches = (
                self._iter_input_batches()
                if self._parquet_reader is None
                else self._parquet_reader._parquet_file.iter_batches(batch_size=min(10000, record_count))
            )
            output_files, total_rows = _stream_split_by_record_count(
                batches,
                schema,
                output_pattern,
                record_count,
                progress_callback=progress_callback,
                compression=(
                    self._parquet_reader._get_compression_type()
                    if self._parquet_reader is not None and output_suffix == ".parquet"
                    else None
                ),
            )
            self.last_split_total_rows = total_rows
            return output_files

        total_rows = self.num_rows
        chunk_sizes = _resolve_split_shape(
            total_rows, file_count=file_count, record_count=record_count
        )
        output_files = _resolve_output_files(output_pattern, len(chunk_sizes))
        if self._parquet_reader is not None:
            result = _split_batches_to_files(
                self._parquet_reader._parquet_file.iter_batches(
                    batch_size=min(10000, max(chunk_sizes))
                ),
                self._parquet_reader.schema,
                output_files,
                chunk_sizes,
                total_rows,
                progress_callback=progress_callback,
                compression=self._parquet_reader._get_compression_type()
                if output_suffix == ".parquet"
                else None,
            )
            self.last_split_total_rows = total_rows
            return result

        result = _split_batches_to_files(
            self._iter_input_batches(),
            self.schema,
            output_files,
            chunk_sizes,
            total_rows,
            progress_callback=progress_callback,
        )
        self.last_split_total_rows = total_rows
        return result

    def convert_file(self, output_path: Union[str, Path], columns: Optional[List[str]] = None) -> int:
        """Convert the input file to a single supported output file."""
        output_path = Path(output_path)
        if self._parquet_reader is not None:
            rows_written = self._parquet_reader.convert_file(output_path, columns=columns)
            self.last_write_total_rows = rows_written
            return rows_written

        rows_written = _write_batches_to_output(
            self._iter_input_batches(columns=columns),
            self._selected_schema(columns),
            output_path,
        )
        self.last_write_total_rows = rows_written
        return rows_written

    def get_stats(self, columns: Optional[List[str]] = None, limit: int = 50) -> List[dict]:
        """Return simple column statistics for the current file."""
        if self._parquet_reader is not None:
            return self._parquet_reader.get_stats(columns=columns, limit=limit)
        selected_schema = self._selected_schema(columns)
        return _compute_stats_from_batches(
            self._iter_input_batches(columns=columns),
            selected_schema,
            limit=limit,
        )


def _reader_batch_iterator(reader: MultiFormatReader, columns: Optional[List[str]] = None) -> Iterator[pa.RecordBatch]:
    """Return a batch iterator for any supported reader."""
    if reader._parquet_reader is not None:
        yield from reader._parquet_reader._parquet_file.iter_batches(columns=columns)
        return
    yield from reader._iter_input_batches(columns=columns)


def _iter_row_dicts(
    reader: MultiFormatReader,
    columns: List[str],
) -> Iterator[dict[str, Any]]:
    """Iterate selected rows as lightweight dicts without materializing full tables."""
    for batch in _reader_batch_iterator(reader, columns=columns):
        batch_dict = batch.to_pydict()
        for row_idx in range(len(batch)):
            yield {column: batch_dict[column][row_idx] for column in columns}


def _build_diff_index(
    reader: MultiFormatReader,
    key_columns: List[str],
    selected_columns: List[str],
    comparable_columns: List[str],
    side_name: str,
) -> dict[tuple[Any, ...], dict[str, Any]]:
    """Build a keyed row index for diffing and reject duplicate keys."""
    index: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in _iter_row_dicts(reader, selected_columns):
        key = tuple(row[column] for column in key_columns)
        if key in index:
            rendered_key = key if len(key) > 1 else key[0]
            raise ValueError(f"Duplicate key found in {side_name} input: {rendered_key}")
        index[key] = {column: row[column] for column in comparable_columns}
    return index


def diff_files(
    left_path: Union[str, Path],
    right_path: Union[str, Path],
    key_columns: List[str],
    columns: Optional[List[str]] = None,
    limit: int = 20,
    summary_only: bool = False,
) -> dict:
    """Compute a keyed diff summary for two parquet/csv files."""
    left_reader = MultiFormatReader(str(left_path))
    right_reader = MultiFormatReader(str(right_path))

    if "xlsx" in {left_reader.input_format, right_reader.input_format}:
        raise ValueError("diff currently supports parquet/csv inputs; convert xlsx first")

    _validate_preview_params(0, left_reader.schema, key_columns)
    _validate_preview_params(0, right_reader.schema, key_columns)
    if columns is not None:
        _validate_preview_params(0, left_reader.schema, columns)
        _validate_preview_params(0, right_reader.schema, columns)

    left_names = left_reader.schema.names
    right_names = right_reader.schema.names
    schema_only_left = [name for name in left_names if name not in right_names]
    schema_only_right = [name for name in right_names if name not in left_names]
    shared_columns = [name for name in left_names if name in right_names]
    schema_type_mismatches = []
    for name in shared_columns:
        left_field = left_reader.schema.field(name)
        right_field = right_reader.schema.field(name)
        if str(left_field.type) != str(right_field.type):
            schema_type_mismatches.append(
                {
                    "column": name,
                    "left_type": str(left_field.type),
                    "right_type": str(right_field.type),
                }
            )

    comparable_columns = columns or [name for name in left_names if name in right_names and name not in key_columns]
    selected_columns = list(dict.fromkeys(key_columns + comparable_columns))
    index_left = left_reader.num_rows <= right_reader.num_rows
    indexed_reader = left_reader if index_left else right_reader
    streamed_reader = right_reader if index_left else left_reader
    indexed_side_name = "left" if index_left else "right"
    streamed_side_name = "right" if index_left else "left"

    indexed_rows = _build_diff_index(
        indexed_reader,
        key_columns,
        selected_columns,
        comparable_columns,
        indexed_side_name,
    )
    missing_from_indexed: List[tuple[Any, ...]] = []
    missing_from_indexed_count = 0
    changed_rows = []
    changed_count = 0
    seen_streamed_keys: set[tuple[Any, ...]] = set()

    for row in _iter_row_dicts(streamed_reader, selected_columns):
        key = tuple(row[column] for column in key_columns)
        if key in seen_streamed_keys:
            rendered_key = key if len(key) > 1 else key[0]
            raise ValueError(f"Duplicate key found in {streamed_side_name} input: {rendered_key}")
        seen_streamed_keys.add(key)

        streamed_values = {column: row[column] for column in comparable_columns}
        if key not in indexed_rows:
            missing_from_indexed_count += 1
            if not summary_only and len(missing_from_indexed) < limit:
                missing_from_indexed.append(key)
            continue

        indexed_values = indexed_rows.pop(key)
        if any(streamed_values[column] != indexed_values[column] for column in comparable_columns):
            changed_count += 1
            if index_left:
                left_values = indexed_values
                right_values = streamed_values
            else:
                left_values = streamed_values
                right_values = indexed_values
            if not summary_only and len(changed_rows) < limit:
                changed_rows.append(
                    {
                        "key": key if len(key) > 1 else key[0],
                        "left": left_values,
                        "right": right_values,
                    }
                )

    remaining_indexed_count = len(indexed_rows)
    if not summary_only:
        remaining_indexed_keys = sorted(indexed_rows.keys())[:limit]
        missing_from_indexed.sort()
        only_left_keys = remaining_indexed_keys if index_left else missing_from_indexed
        only_right_keys = missing_from_indexed if index_left else remaining_indexed_keys
    else:
        only_left_keys = []
        only_right_keys = []

    only_left_count = remaining_indexed_count if index_left else missing_from_indexed_count
    only_right_count = missing_from_indexed_count if index_left else remaining_indexed_count

    return {
        "row_count_delta": left_reader.num_rows - right_reader.num_rows,
        "only_left_count": only_left_count,
        "only_right_count": only_right_count,
        "changed_count": changed_count,
        "schema_only_left": [] if summary_only else [{"column": name} for name in schema_only_left[:limit]],
        "schema_only_right": [] if summary_only else [{"column": name} for name in schema_only_right[:limit]],
        "schema_type_mismatches": [] if summary_only else schema_type_mismatches[:limit],
        "only_left": [{"key": key if len(key) > 1 else key[0]} for key in only_left_keys[:limit]],
        "only_right": [{"key": key if len(key) > 1 else key[0]} for key in only_right_keys[:limit]],
        "changed_rows": [] if summary_only else changed_rows[:limit],
    }


def merge_files(input_paths: List[Union[str, Path]], output_path: Union[str, Path]) -> int:
    """Merge multiple compatible inputs into a single output file."""
    if len(input_paths) == 0:
        raise ValueError("At least one input file must be provided")

    output_path = Path(output_path)
    if output_path.exists():
        raise FileExistsError(f"Output file already exists: {output_path}")

    readers = [MultiFormatReader(str(path)) for path in input_paths]
    schemas = [reader.schema for reader in readers]
    try:
        unified_schema = pa.unify_schemas(schemas)
    except (pa.ArrowInvalid, pa.ArrowTypeError) as e:
        raise ValueError(f"Incompatible schemas: {e}") from e

    writer = _open_chunk_writer(output_path, unified_schema, compression="SNAPPY" if output_path.suffix.lower() == ".parquet" else None)
    total_rows = 0
    try:
        for reader in readers:
            for batch in _reader_batch_iterator(reader):
                batch_table = pa.Table.from_batches([batch], schema=batch.schema)
                if batch_table.schema != unified_schema:
                    batch = batch_table.cast(unified_schema).to_batches()[0]
                writer.write_batch(batch)
                total_rows += len(batch)
    except Exception:
        try:
            writer.close()
        except Exception:
            pass
        try:
            output_path.unlink()
        except FileNotFoundError:
            pass
        except OSError:
            pass
        raise

    writer.close()
    return total_rows
